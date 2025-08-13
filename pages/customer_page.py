import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from util import dbutil
from util.utils import center_window
import os

def CustomerPage(parent, username):
    frame = ttk.Frame(parent)
    # 工具栏
    toolbar = ttk.Frame(frame)
    toolbar.pack(fill=tk.X, pady=5)
    # 左侧图标按钮
    btn_add = ttk.Button(toolbar, text="➕", width=3)
    btn_add.pack(side=tk.LEFT, padx=3)
    btn_edit = ttk.Button(toolbar, text="✏️", width=3)
    btn_edit.pack(side=tk.LEFT, padx=3)
    btn_del = ttk.Button(toolbar, text="🗑️", width=3)
    btn_del.pack(side=tk.LEFT, padx=3)
    btn_import = ttk.Button(toolbar, text="📤", width=3)
    btn_import.pack(side=tk.LEFT, padx=3)
    # 搜索框 - 放到导入按钮右侧
    search_frame = ttk.Frame(toolbar)
    search_frame.pack(side=tk.LEFT, padx=10)
    search_var = tk.StringVar()
    search_entry = ttk.Entry(search_frame, textvariable=search_var, width=20)
    search_entry.pack(side=tk.LEFT, padx=5)
    
    def on_search():
        keyword = search_var.get().strip().lower()
        if not keyword:
            load_data()
            return
        # 清空当前数据
        for row in tree.get_children():
            tree.delete(row)
        # 搜索并加载数据
        for idx, row in enumerate(dbutil.get_all_customers(), 1):
            # row: (id, name, address, phone, logistics_info)
            name = row[1].lower()
            phone = row[3].lower()
            if keyword in name or keyword in phone:
                tree.insert("", tk.END, values=(idx, *row[1:]))
    
    btn_search = ttk.Button(search_frame, text="搜索", command=on_search)
    btn_search.pack(side=tk.LEFT)
    
    # 右侧密码设置按钮
    right_toolbar = ttk.Frame(toolbar)
    right_toolbar.pack(side=tk.RIGHT, padx=10)
    btn_password = ttk.Button(right_toolbar, text="密码设置", width=8)
    btn_password.pack(side=tk.RIGHT)

    # 为搜索框添加回车事件
    search_entry.bind('<Return>', lambda event: on_search())

    # 密码设置功能实现
    def handle_password_setting():
        from util.dbutil import get_user_unipassword_by_username, update_user_unipassword
        current_password = get_user_unipassword_by_username(username)

        if not current_password or current_password.strip() == '':
            # 设置新密码
            def set_new_password():
                new_pwd = pwd_entry.get()
                confirm_pwd = confirm_entry.get()
                # 允许密码为空
                new_pwd = new_pwd.strip()
                if new_pwd != confirm_pwd:
                    messagebox.showwarning("提示", "两次输入的密码不一致！")
                    return
                update_user_unipassword(username, new_pwd)
                dialog.destroy()
                messagebox.showinfo("成功", "密码设置成功！")

            dialog = tk.Toplevel(frame)
            dialog.title("设置密码")
            dialog.transient(frame)
            dialog.grab_set()
            center_window(dialog, 310, 220)

            ttk.Label(dialog, text="设置新密码:", font=("微软雅黑", 10)).grid(row=0, column=0, sticky=tk.W, padx=10, pady=10)
            pwd_entry = ttk.Entry(dialog, show="*", width=20)
            pwd_entry.grid(row=0, column=1, padx=5)

            ttk.Label(dialog, text="确认密码:", font=("微软雅黑", 10)).grid(row=1, column=0, sticky=tk.W, padx=10, pady=10)
            confirm_entry = ttk.Entry(dialog, show="*", width=20)
            confirm_entry.grid(row=1, column=1, padx=5)

            # 添加提示信息
            ttk.Label(dialog, text="提示: 新密码为空则关闭密码保护", font=('微软雅黑', 9, 'italic'), foreground='gray').grid(row=2, column=0, columnspan=2, padx=10, pady=(5, 10), sticky=tk.W)
            btn_frame = ttk.Frame(dialog)
            btn_frame.grid(row=3, column=0, columnspan=2, pady=5)
            ttk.Button(btn_frame, text="确定", command=set_new_password).pack(side=tk.LEFT, padx=10)
            ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side=tk.LEFT)

        else:
            # 修改密码
            def change_password():
                old_pwd = old_pwd_entry.get()
                new_pwd = new_pwd_entry.get()
                confirm_pwd = confirm_pwd_entry.get()

                if old_pwd != current_password:
                    messagebox.showwarning("提示", "原密码输入错误！")
                    return
                # 允许密码为空
                new_pwd = new_pwd.strip()
                if new_pwd != confirm_pwd:
                    messagebox.showwarning("提示", "两次输入的新密码不一致！")
                    return
                update_user_unipassword(username, new_pwd)
                dialog.destroy()
                messagebox.showinfo("成功", "密码修改成功！")

            dialog = tk.Toplevel(frame)
            dialog.title("修改密码")
            dialog.transient(frame)
            dialog.grab_set()
            center_window(dialog, 300, 180)

            ttk.Label(dialog, text="原密码:", font=("微软雅黑", 10)).grid(row=0, column=0, sticky=tk.W, padx=10, pady=8)
            old_pwd_entry = ttk.Entry(dialog, show="*", width=20)
            old_pwd_entry.grid(row=0, column=1, padx=5)

            ttk.Label(dialog, text="新密码:", font=("微软雅黑", 10)).grid(row=1, column=0, sticky=tk.W, padx=10, pady=8)
            new_pwd_entry = ttk.Entry(dialog, show="*", width=20)
            new_pwd_entry.grid(row=1, column=1, padx=5)

            ttk.Label(dialog, text="确认新密码:", font=("微软雅黑", 10)).grid(row=2, column=0, sticky=tk.W, padx=10, pady=8)
            confirm_pwd_entry = ttk.Entry(dialog, show="*", width=20)
            confirm_pwd_entry.grid(row=2, column=1, padx=5)

            # 添加提示信息
            ttk.Label(dialog, text="提示: 新密码为空则关闭密码保护", font=('微软雅黑', 9, 'italic'), foreground='gray').grid(row=3, column=0, columnspan=2, padx=10, pady=(5, 10), sticky=tk.W)
            btn_frame = ttk.Frame(dialog)
            btn_frame.grid(row=4, column=0, columnspan=2, pady=5)
            ttk.Button(btn_frame, text="确定", command=change_password).pack(side=tk.LEFT, padx=10)
            ttk.Button(btn_frame, text="取消", command=dialog.destroy).pack(side=tk.LEFT)

    # tooltip
    def add_tooltip(widget, text):
        tip = tk.Toplevel(widget)
        tip.withdraw()
        tip.overrideredirect(True)
        label = tk.Label(tip, text=text, background="#ffffe0", relief=tk.SOLID, borderwidth=1, font=("微软雅黑", 9))
        label.pack(ipadx=4)
        def enter(event):
            x, y, cx, cy = widget.bbox("insert") if hasattr(widget, 'bbox') else (0, 0, 0, 0)
            x += widget.winfo_rootx() + 30
            y += widget.winfo_rooty() + 20
            tip.geometry(f"+{x}+{y}")
            tip.deiconify()
        def leave(event):
            tip.withdraw()
        widget.bind('<Enter>', enter)
        widget.bind('<Leave>', leave)
    # 为按钮添加提示
    add_tooltip(btn_add, "新增客户")
    add_tooltip(btn_edit, "修改客户信息")
    add_tooltip(btn_del, "删除客户信息")
    add_tooltip(btn_import, "批量导入客户")
    add_tooltip(btn_password, "设置/修改密码")
    btn_password.config(command=handle_password_setting)

    # 批量导入客户
    def batch_import_customers():
        # 创建批量导入对话框
        dialog = tk.Toplevel(frame)
        dialog.title("批量导入客户")
        dialog.transient(frame)
        dialog.grab_set()
        center_window(dialog, 400, 250)

        # 对话框内容
        ttk.Label(dialog, text="请先下载Excel模板，填写客户信息后再导入", font=('微软雅黑', 10)).pack(pady=15)

        # 模板下载按钮
        def download_template():
            try:
                # 创建Excel模板
                from openpyxl import Workbook
                
                # 定义模板数据
                template_data = {
                    '姓名': ['客户姓名1', '客户姓名2'],
                    '地址': ['客户地址1', '客户地址2'],
                    '电话': ['13800138000', '13900139000'],
                    '物流信息': ['物流信息1', '物流信息2']
                }
                
                # 使用openpyxl创建工作簿
                wb = Workbook()
                ws = wb.active
                ws.title = '客户信息'
                
                # 添加表头
                headers = list(template_data.keys())
                ws.append(headers)
                
                # 添加示例数据
                for i in range(len(template_data['姓名'])):
                    row_data = [template_data[col][i] for col in headers]
                    ws.append(row_data)
                
                # 保存文件
                file_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx")],
                    title="保存模板文件",
                    initialfile="客户信息模板"
                )
                if file_path:
                    wb.save(file_path)
                    messagebox.showinfo("成功", f"模板已成功下载到: {file_path}")
            except Exception as e:
                messagebox.showerror("错误", f"下载模板失败: {str(e)}")

        btn_download = ttk.Button(dialog, text="下载Excel模板", command=download_template)
        btn_download.pack(pady=10)

        # 文件选择按钮
        def select_file():
            file_path = filedialog.askopenfilename(
                title="选择Excel文件",
                filetypes=[("Excel files", "*.xlsx *.xls")]
            )
            if not file_path:
                return

            try:
                # 读取Excel文件
                from openpyxl import load_workbook
                
                wb = load_workbook(file_path, read_only=True)
                ws = wb.active
                
                # 获取表头
                headers = []
                for cell in ws[1]:
                    headers.append(cell.value)
                
                # 检查必要的列是否存在
                required_columns = ['姓名', '地址', '电话']
                for col in required_columns:
                    if col not in headers:
                        messagebox.showerror("错误", f"Excel文件中缺少必要的列: {col}")
                        return

                # 导入数据
                success_count = 0
                # 跳过表头行，从第二行开始读取
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = {headers[i]: row[i] for i in range(len(headers))}
                    
                    name = str(row_dict['姓名']).strip() if row_dict['姓名'] is not None else ''
                    address = str(row_dict['地址']).strip() if row_dict['地址'] is not None else ''
                    phone = str(row_dict['电话']).strip() if row_dict['电话'] is not None else ''
                    
                    # 处理物流信息
                    logistics_info = ''
                    if '物流信息' in headers:
                        logistics_info = str(row_dict['物流信息']).strip() if row_dict['物流信息'] is not None else ''

                    if name:
                        dbutil.insert_customer(name, address, phone, logistics_info)
                        success_count += 1

                load_data()
                messagebox.showinfo("成功", f"成功导入 {success_count} 条客户信息！")
                dialog.destroy()
            except Exception as e:
                messagebox.showerror("错误", f"导入失败: {str(e)}")

        btn_select = ttk.Button(dialog, text="选择Excel文件", command=select_file)
        btn_select.pack(pady=10)

        # 取消按钮
        btn_cancel = ttk.Button(dialog, text="取消", command=dialog.destroy)
        btn_cancel.pack(pady=10)

    btn_import.config(command=batch_import_customers)

    # 确保openpyxl已安装
    try:
        import openpyxl
    except ImportError:
        # 如果未安装，提供安装提示
        def prompt_install_openpyxl():
            if messagebox.askyesno("提示", "批量导入功能需要openpyxl库。是否现在安装?"):
                try:
                    import subprocess
                    subprocess.call(["pip", "install", "openpyxl"])
                    messagebox.showinfo("成功", "openpyxl库安装成功！")
                except Exception as e:
                    messagebox.showerror("错误", f"安装失败: {str(e)}")
        btn_import.config(command=prompt_install_openpyxl)
    # 表格
    columns = ("no", "name", "address", "phone", "logistics_info")
    headers = [
        ("no", "序号"),
        ("name", "姓名"),
        ("address", "地址"),
        ("phone", "电话"),
        ("logistics_info", "物流信息")
    ]
    tree = ttk.Treeview(frame, columns=columns, show="headings", height=15)
    tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
    for col, text in headers:
        if col == "logistics_info":
            tree.heading(col, text=text)
            tree.column(col, anchor=tk.CENTER, width=310)
        else:
            tree.heading(col, text=text)
            tree.column(col, anchor=tk.CENTER, width=30)
    def load_data():
        for row in tree.get_children():
            tree.delete(row)
        for idx, row in enumerate(dbutil.get_all_customers(), 1):
            # row: (id, name, address, phone, logistics_info)
            tree.insert("", tk.END, values=(idx, *row[1:]))
    frame.refresh = load_data
    load_data()
    # 新增客户
    def open_add_dialog():
        dialog = tk.Toplevel(frame)
        dialog.title("新增客户")
        dialog.transient(frame)
        dialog.grab_set()
        center_window(dialog, 400, 320)
        fields = [
            ("姓名", "name"),
            ("地址", "address"),
            ("电话", "phone")
        ]
        vars = {}
        for idx, (label, key) in enumerate(fields):
            ttk.Label(dialog, text=label+":").grid(row=idx, column=0, sticky=tk.W, padx=10, pady=6)
            var = tk.StringVar()
            ttk.Entry(dialog, textvariable=var, width=30).grid(row=idx, column=1, padx=5)
            vars[key] = var
        # 物流信息动态输入
        logistics_entries = []
        def add_logistics_entry(default_val=""):
            row = len(fields) + len(logistics_entries)
            entry_var = tk.StringVar(value=default_val)
            entry = ttk.Entry(dialog, textvariable=entry_var, width=30)
            entry.grid(row=row, column=1, padx=(5,0), pady=2, sticky=tk.W)
            logistics_entries.append(entry_var)
        # 初始一个物流信息输入框
        ttk.Label(dialog, text="物流信息:").grid(row=len(fields), column=0, sticky=tk.W, padx=10, pady=6)
        add_logistics_entry()
        def on_add_logistics():
            add_logistics_entry()
        btn_add_logistics = ttk.Button(dialog, text="➕", width=2, command=on_add_logistics)
        btn_add_logistics.grid(row=len(fields), column=2, padx=(2,10), sticky=tk.W)
        def on_ok():
            if not vars['name'].get().strip():
                messagebox.showwarning("提示", "姓名不能为空！")
                return
            logistics_info = ";".join([v.get().strip() for v in logistics_entries if v.get().strip()])
            dbutil.insert_customer(
                vars['name'].get().strip(),
                vars['address'].get().strip(),
                vars['phone'].get().strip(),
                logistics_info
            )
            dialog.destroy()
            load_data()
            messagebox.showinfo("成功", "新增客户成功！")
        ttk.Button(dialog, text="确定", command=on_ok, width=12).grid(row=6+len(logistics_entries), column=0, columnspan=3, pady=12)
        dialog.wait_window()
    btn_add.config(command=open_add_dialog)
    # 修改客户
    def open_edit_dialog():
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要修改的客户！")
            return
        item = tree.item(selected[0])
        values = item['values']
        dialog = tk.Toplevel(frame)
        dialog.title("修改客户信息")
        dialog.transient(frame)
        dialog.grab_set()
        center_window(dialog, 400, 320)
        fields = [
            ("姓名", "name"),
            ("地址", "address"),
            ("电话", "phone")
        ]
        vars = {}
        for idx, (label, key) in enumerate(fields):
            ttk.Label(dialog, text=label+":").grid(row=idx, column=0, sticky=tk.W, padx=10, pady=6)
            var = tk.StringVar(value=values[idx+1])
            ttk.Entry(dialog, textvariable=var, width=30).grid(row=idx, column=1, padx=5)
            vars[key] = var
        # 物流信息动态输入
        logistics_entries = []
        logistics_list = values[4].split(';') if values[4] else ['']
        ttk.Label(dialog, text="物流信息:").grid(row=len(fields), column=0, sticky=tk.W, padx=10, pady=6)
        def add_logistics_entry_edit(default_val=""):
            row = len(fields) + len(logistics_entries)
            entry_var = tk.StringVar(value=default_val)
            entry = ttk.Entry(dialog, textvariable=entry_var, width=30)
            entry.grid(row=row, column=1, padx=(5,0), pady=2, sticky=tk.W)
            logistics_entries.append(entry_var)
        for l in logistics_list:
            add_logistics_entry_edit(l)
        def on_add_logistics():
            add_logistics_entry_edit()
        btn_add_logistics = ttk.Button(dialog, text="➕", width=2, command=on_add_logistics)
        btn_add_logistics.grid(row=len(fields), column=2, padx=(2,10), sticky=tk.W)
        def on_ok():
            if not vars['name'].get().strip():
                messagebox.showwarning("提示", "姓名不能为空！")
                return
            logistics_info = ";".join([v.get().strip() for v in logistics_entries if v.get().strip()])
            # 用 tree.item 的 iid 获取真实主键
            cid = tree.item(selected[0], 'values')[0]
            all_rows = dbutil.get_all_customers()
            for idx, row in enumerate(all_rows, 1):
                if str(idx) == str(cid):
                    real_id = row[0]
                    break
            else:
                real_id = cid
            dbutil.update_customer(
                real_id,
                vars['name'].get().strip(),
                vars['address'].get().strip(),
                vars['phone'].get().strip(),
                logistics_info
            )
            dialog.destroy()
            load_data()
            messagebox.showinfo("成功", "修改成功！")
        ttk.Button(dialog, text="确定", command=on_ok, width=12).grid(row=6+len(logistics_entries), column=0, columnspan=3, pady=12)
        dialog.wait_window()
    btn_edit.config(command=open_edit_dialog)
    # 删除客户
    def on_delete():
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要删除的客户！")
            return
        item = tree.item(selected[0])
        cid = item['values'][0]
        # 查找真实主键id
        all_rows = dbutil.get_all_customers()
        for idx, row in enumerate(all_rows, 1):
            if str(idx) == str(cid):
                real_id = row[0]
                break
        else:
            real_id = cid
        if not messagebox.askyesno("确认", "确定要删除该客户信息吗？"):
            return
        dbutil.delete_customer(real_id)
        load_data()
        messagebox.showinfo("成功", "删除成功！")
    btn_del.config(command=on_delete)
    return frame
