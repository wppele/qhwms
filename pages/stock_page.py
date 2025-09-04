# 采购入库页面
import tkinter as tk
from tkinter import ttk, filedialog
import tkinter.messagebox as messagebox
from util.utils import center_window
from util import dbutil
import util.utils as utils

def StockPage(parent, main_win):
    frame = ttk.Frame(parent)
    # 工具栏（包含按钮和搜索控件）
    stock_toolbar = ttk.Frame(frame)
    stock_toolbar.pack(fill=tk.X, padx=10, pady=10, anchor=tk.N)
    # 左侧按钮
    def open_add_stock_dialog_and_refresh():
        open_add_stock_dialog()
        load_stock_data()
    # 定义add_tooltip函数
    def add_tooltip(widget, text):
        tip = tk.Toplevel(widget)
        tip.withdraw()
        tip.overrideredirect(True)
        label = tk.Label(tip, text=text, background="#ffffe0", relief=tk.SOLID, borderwidth=1, font=("微软雅黑", 9))
        label.pack(ipadx=4)
        def enter(event):
            x, y, cx, cy = widget.bbox("insert") if hasattr(widget, 'bbox') else (0, 0, 0, 0)
            x += widget.winfo_rootx() + 30
            y += widget.winfo_rooty() + 20
            tip.geometry(f"+{x}+{y}")
            tip.deiconify()
        def leave(event):
            tip.withdraw()
        widget.bind('<Enter>', enter)
        widget.bind('<Leave>', leave)
    
    # 左侧按钮
    def open_add_stock_dialog_and_refresh():
        open_add_stock_dialog()
        load_stock_data()
    btn_add = ttk.Button(stock_toolbar, text="➕", width=3, command=open_add_stock_dialog_and_refresh)
    btn_add.pack(side=tk.LEFT, padx=3)
    add_tooltip(btn_add, "新增库存")
    
    # 批量导入按钮
    def handle_batch_import():
        # 创建批量导入对话框
        import_dialog = tk.Toplevel(main_win)
        import_dialog.title("批量导入库存")
        import_dialog.geometry("500x300")
        center_window(import_dialog)
        
        # 设置对话框模态
        import_dialog.grab_set()
        
        # 导入按钮
        def do_import():
            file_path = file_path_var.get().strip()
            if not file_path:
                messagebox.showerror("错误", "请选择文件")
                return
            
            try:
                # 检查是否安装了openpyxl
                import openpyxl
            except ImportError:
                messagebox.showerror("错误", "请安装openpyxl库\n命令: pip install openpyxl")
                return
            
            try:
                # 读取Excel文件
                from openpyxl import load_workbook
                
                wb = load_workbook(file_path, read_only=True)
                ws = wb.active
                
                # 获取表头
                headers = []
                for cell in ws[1]:
                    headers.append(cell.value)
                
                # 检查必要的列是否存在
                required_columns = ['厂家', '货号', '数量', '单价', '入库日期']
                missing_columns = [col for col in required_columns if col not in headers]
                if missing_columns:
                    messagebox.showerror("错误", f"Excel文件缺少必要的列: {', '.join(missing_columns)}")
                    return
                
                # 准备库存数据列表
                stocks = []
                # 跳过表头行，从第二行开始读取
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = {headers[i]: row[i] for i in range(len(headers))}
                    
                    factory = str(row_dict['厂家']).strip() if row_dict['厂家'] is not None else ''
                    product_no = str(row_dict['货号']).strip() if row_dict['货号'] is not None else ''
                    size = str(row_dict.get('尺码', '')).strip() if row_dict.get('尺码') is not None else ''
                    color = str(row_dict.get('颜色', '')).strip() if row_dict.get('颜色') is not None else ''
                    unit = str(row_dict.get('单位', '个')).strip() if row_dict.get('单位') is not None else '个'
                    in_quantity = row_dict['数量']
                    price = row_dict['单价']
                    in_date = str(row_dict['入库日期']).strip() if row_dict['入库日期'] is not None else ''
                    
                    # 验证数据
                    if not all([factory, product_no, in_quantity, price, in_date]):
                        continue
                    
                    try:
                        in_quantity = int(in_quantity)
                        price = float(price)
                        total = in_quantity * price
                    except ValueError:
                        continue
                    
                    stocks.append((factory, product_no, size, color, unit, in_quantity, price, total, in_date))
                
                if not stocks:
                    messagebox.showerror("错误", "没有有效的库存数据")
                    return
                
                # 批量插入库存
                inserted_count = dbutil.batch_insert_stocks(stocks)
                
                messagebox.showinfo("成功", f"成功导入 {inserted_count} 条库存记录")
                import_dialog.destroy()
                load_stock_data()
            except Exception as e:
                messagebox.showerror("错误", f"导入失败: {str(e)}")
        
        # 下载模板按钮
        def download_template():
            try:
                # 检查是否安装了openpyxl
                import openpyxl
            except ImportError:
                messagebox.showerror("错误", "请安装openpyxl库\n命令: pip install openpyxl")
                return
            
            try:
                # 创建模板数据
                from openpyxl import Workbook
                
                # 创建工作簿和工作表
                wb = Workbook()
                ws = wb.active
                ws.title = '库存模板'
                
                # 添加表头
                headers = ['厂家', '货号', '尺码', '颜色', '单位', '数量', '单价', '入库日期']
                ws.append(headers)
                
                # 添加示例数据
                current_date = utils.get_current_date()
                example_data = ['示例厂家', '示例货号', '示例尺码', '示例颜色', '个', 10, 99.99, current_date]
                ws.append(example_data)
                
                # 让用户选择保存路径
                save_path = filedialog.asksaveasfilename(
                    defaultextension='.xlsx',
                    filetypes=[("Excel files", "*.xlsx")],
                    title="保存模板",
                    initialfile="采购模板"
                )
                
                if save_path:
                    wb.save(save_path)
                    messagebox.showinfo("成功", f"模板已保存到: {save_path}")
            except Exception as e:
                messagebox.showerror("错误", f"保存模板失败: {str(e)}")
        
        btn_template = ttk.Button(import_dialog, text="下载模板", command=download_template)
        btn_template.pack(pady=10)
        
        # 文件选择框架
        file_frame = ttk.Frame(import_dialog)
        file_frame.pack(fill=tk.X, padx=20, pady=10)
        
        file_path_var = tk.StringVar()
        ttk.Entry(file_frame, textvariable=file_path_var, width=15).pack(side=tk.LEFT, padx=5)
        
        def select_file():
            file_path = filedialog.askopenfilename(
                filetypes=[("Excel files", "*.xlsx *.xls")]
            )
            if file_path:
                file_path_var.set(file_path)
        
        ttk.Button(file_frame, text="浏览", command=select_file).pack(side=tk.LEFT, padx=5)
        
        # 导入按钮
        btn_import = ttk.Button(import_dialog, text="导入", command=do_import)
        btn_import.pack(pady=10)
        
        # 居中显示
        center_window(import_dialog)
    
# 批量导入按钮移至修改按钮后
    
    # 其他按钮和工具栏内容
    # ...existing code...
    # 统一的库存提交方法，提升到StockPage作用域
    def handle_stock_submit(vars, entry_refs, error_label, dialog, is_edit=False, stock_id=None):
        for k, v in vars.items():
            if k != 'total' and not v.get().strip():
                error_label['text'] = "所有字段不能为空！"
                return
        try:
            vars["total"].set(utils.calculate_total(vars["in_quantity"].get(), vars["price"].get()))
            factory = vars['factory'].get().strip()
            product_no = vars['product_no'].get().strip()
            size = vars['size'].get().strip()
            color = vars['color'].get().strip()
            unit = vars.get('unit', tk.StringVar()).get().strip() if 'unit' in vars else ''
            in_quantity = int(vars['in_quantity'].get())
            price = float(vars['price'].get())
            total = float(vars['total'].get())
            in_date = vars.get('in_date', tk.StringVar()).get().strip() if 'in_date' in vars else utils.get_current_date()
            if is_edit:
                dbutil.update_stock_by_id(
                    stock_id,
                    factory,
                    product_no,
                    size,
                    color,
                    unit,
                    in_quantity,
                    price,
                    total
                )
                dbutil.update_inventory_by_stock_id(
                    stock_id,
                    factory,
                    product_no,
                    size,
                    color,
                    unit,
                    in_quantity
                )
                tk.messagebox.showinfo("成功", "修改成功！")
                dialog.destroy()  # 关闭对话框
            else:
                dbutil.insert_stock(
                    factory,
                    product_no,
                    size,
                    color,
                    unit,
                    in_quantity,
                    price,
                    total,
                    in_date
                )
                stock_rows = dbutil.get_all_stock()
                if stock_rows:
                    stock_id = stock_rows[0][0]
                # 处理库存合并和日志
                found = None
                inventory_rows = dbutil.get_all_inventory()
                for row in inventory_rows:
                    if row[2] == factory and row[3] == product_no and row[5] == color:
                        found = row
                        break
                if found:
                    inventory_id = found[0]
                    old_size = found[4] or ''
                    new_size = size or ''
                    merged_size = dbutil.merge_size(old_size, new_size) if hasattr(dbutil, 'merge_size') else ','.join(sorted(set(s.strip() for s in (old_size + ',' + new_size).split(',') if s.strip())))
                    dbutil.update_inventory_size_by_id(inventory_id, merged_size)
                    dbutil.decrease_inventory_by_id(inventory_id, -in_quantity)
                    dbutil.insert_stock_log(
                        factory,
                        product_no,
                        merged_size,
                        color,
                        in_quantity,
                        '入库',
                        in_date
                    )
                    if hasattr(main_win, 'refresh_logs'):
                        main_win.refresh_logs()
                    tk.messagebox.showinfo("成功", f"已存在相同库存（尺码已合并为：{merged_size}），数量已增加 {in_quantity}！")
                else:
                    dbutil.insert_inventory_from_stock(
                        stock_id,
                        factory,
                        product_no,
                        size,
                        color,
                        unit,
                        in_quantity
                    )
                    dbutil.insert_stock_log(
                        factory,
                        product_no,
                        size,
                        color,
                        in_quantity,
                        '入库',
                        in_date
                    )
                    if hasattr(main_win, 'refresh_logs'):
                        main_win.refresh_logs()
                    tk.messagebox.showinfo("成功", "新增成功！")
                vars['color'].set("")
                vars['in_quantity'].set("")
                vars['price'].set("")
                vars['total'].set("")
                entry_refs['color'].focus_set()
            load_stock_data()
            # 新增成功后不关闭对话框，仅清空颜色、数量、合计，其他字段保留
            vars['color'].set("")
            vars['in_quantity'].set("")
            vars['total'].set("")
            entry_refs['color'].focus_set()
        except Exception as e:
            error_label['text'] = str(e)

    # 右侧搜索控件（修正位置和作用域）
    search_frame = ttk.Frame(stock_toolbar)
    search_frame.pack(side=tk.RIGHT, padx=0)
    ttk.Label(search_frame, text="厂家:").pack(side=tk.LEFT)
    search_factory = tk.StringVar()
    ttk.Entry(search_frame, textvariable=search_factory, width=10).pack(side=tk.LEFT, padx=3)
    ttk.Label(search_frame, text="货号:").pack(side=tk.LEFT)
    search_product_no = tk.StringVar()
    ttk.Entry(search_frame, textvariable=search_product_no, width=10).pack(side=tk.LEFT, padx=3)
    
    # 添加日期选择条件框
    ttk.Label(search_frame, text="日期:").pack(side=tk.LEFT)
    search_date = tk.StringVar()
    try:
        from tkcalendar import DateEntry
        has_tkcalendar = True
    except ImportError:
        has_tkcalendar = False
    
    # 清空日期函数
    def clear_date():
        search_date.set("")
        if has_tkcalendar:
            date_entry.delete(0, tk.END)
    
    if has_tkcalendar:
        date_entry = DateEntry(search_frame, textvariable=search_date, width=10, date_pattern='yyyy-mm-dd')
        date_entry.pack(side=tk.LEFT, padx=3)
        date_entry.delete(0, tk.END)  # 清空默认日期
        ttk.Button(search_frame, text="清空", command=clear_date, width=4).pack(side=tk.LEFT, padx=1)
    else:
        ttk.Entry(search_frame, textvariable=search_date, width=12).pack(side=tk.LEFT, padx=3)
        ttk.Label(search_frame, text="(格式: yyyy-mm-dd)", font=('Arial', 8)).pack(side=tk.LEFT)
        ttk.Button(search_frame, text="清空", command=clear_date, width=4).pack(side=tk.LEFT, padx=1)
    ttk.Label(search_frame, text="是否结账:").pack(side=tk.LEFT)
    search_settled = tk.StringVar(value="全部")
    ttk.Combobox(search_frame, textvariable=search_settled, values=["全部", "是", "否"], width=5, state="readonly").pack(side=tk.LEFT, padx=3)
    def do_search():
        factory = search_factory.get().strip()
        product_no = search_product_no.get().strip()
        search_date_str = search_date.get().strip()
        settled = search_settled.get()
        results = []
        
        import datetime
        search_date_obj = None
        if search_date_str:
            try:
                search_date_obj = datetime.datetime.strptime(search_date_str, "%Y-%m-%d").date()
            except ValueError:
                pass
        
        for row in dbutil.get_all_stock():
            # id, factory, product_no, size, color,unit, in_quantity, price, total, is_settled, in_date
            _, f, p, _, _, _, _, _, _, s, in_date = row
            if factory and factory not in f:
                continue
            if product_no and product_no not in p:
                continue
            
            # 日期过滤
            if search_date_obj:
                try:
                    row_date = datetime.datetime.strptime(in_date, "%Y-%m-%d").date()
                    if row_date != search_date_obj:
                        continue
                except ValueError:
                    continue
            
            if settled != "全部":
                if (settled == "是" and not s) or (settled == "否" and s):
                    continue
            results.append(row)
        # 刷新表格
        for r in tree.get_children():
            tree.delete(r)
        for idx, row in enumerate(results, 1):
            id_, factory, product_no, size, color,unit, in_quantity, price, total, is_settled, in_date = row
            is_settled = "是" if is_settled else "否"
            tree.insert("", tk.END, iid=str(id_), values=[
                idx, 
                factory, 
                product_no, 
                size, 
                color, 
                unit,
                in_quantity,  # 入库数量
                price,  # 单价
                total, 
                is_settled,
                in_date
            ])
    search_btn = ttk.Button(search_frame, text="搜索", command=do_search, width=8)
    search_btn.pack(side=tk.LEFT, padx=6)
    
    # 绑定回车键到搜索按钮
    search_factory_entry = search_frame.winfo_children()[1]  # 厂家输入框
    search_product_no_entry = search_frame.winfo_children()[3]  # 货号输入框
    
    def on_enter_pressed(event):
        do_search()
    
    search_factory_entry.bind('<Return>', on_enter_pressed)
    search_product_no_entry.bind('<Return>', on_enter_pressed)
    
    # 如果有日期输入框，也绑定回车键
    try:
        date_entry.bind('<Return>', on_enter_pressed)
    except NameError:
        # 如果没有tkcalendar，日期输入框是普通Entry
        date_entry_widget = search_frame.winfo_children()[7]  # 日期输入框
        date_entry_widget.bind('<Return>', on_enter_pressed)
    # 表格区，隐藏id，新增序号列，去掉可用数量
    columns = ("no", "factory", "product_no", "size", "color", "unit", "in_quantity", "price", "total", "is_settled", "in_date")
    headers = [
        ("no", "序号"),
        ("factory", "厂家"),
        ("product_no", "货号"),
        ("size", "尺码"),
        ("color", "颜色"),
        ("unit", "单位"),
        ("in_quantity", "入库数量"),
        ("price", "单价"),
        ("total", "合计"),
        ("is_settled", "是否结账"),
        ("in_date", "入库时间")
    ]
    tree = ttk.Treeview(frame, columns=columns, show="headings", height=12, selectmode="extended")
    # 添加垂直滚动条
    vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=vsb.set)
    tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10,0), pady=10)
    vsb.pack(side=tk.LEFT, fill=tk.Y, pady=10)
    # 双击结账
    def on_settle_selected(event=None):
        row_id = tree.focus() if event is None else tree.identify_row(event.y)
        if not row_id:
            tk.messagebox.showwarning("提示", "请先选择要结账的库存记录！")
            return
        tree.selection_set(row_id)
        selected = tree.selection()
        # 创建日期选择对话框
        date_dialog = tk.Toplevel(main_win)
        date_dialog.title("选择结账日期")
        date_dialog.transient(main_win)
        date_dialog.grab_set()
        center_window(date_dialog, 300, 120)
        ttk.Label(date_dialog, text="请选择结账日期:").pack(pady=10)
        date_var = tk.StringVar()
        date_entry = ttk.Entry(date_dialog, textvariable=date_var)
        date_entry.pack(pady=5)
        date_entry.insert(0, utils.get_current_date())  # 默认当前日期
        def confirm_settle():
            settle_date = date_var.get()
            if not settle_date:
                tk.messagebox.showwarning("提示", "请选择结账日期！")
                return
            # 提前筛选出未结账的物品，避免不必要的循环处理
            unsettled_items = []
            settled_items = []
            for iid in selected:
                item = tree.item(iid)
                values = item['values']
                if values[9] == "是":
                    settled_items.append(values)
                else:
                    unsettled_items.append((iid, item))

            # 记录成功结账和已结账的物品
            # 使用提前筛选的列表进行处理
            settled_count = 0
            already_settled_count = len(settled_items)
            
            for iid, item in unsettled_items:
                values = item['values']
                # 执行结账
                dbutil.settle_stock_by_id(iid)
                # 写入结账日志
                dbutil.insert_settle_log(
                    values[1],  # factory
                    values[2],  # product_no
                    values[3],  # size
                    values[4],  # color
                    int(values[6]),  # in_quantity
                    float(values[7]),  # price
                    float(values[8]),  # total
                    settle_date
                )
                settled_count += 1
            # 自动刷新日志页面
            if hasattr(main_win, 'refresh_logs'):
                main_win.refresh_logs()
            load_stock_data()
            date_dialog.destroy()
            message = f"成功结账 {settled_count} 条记录！\n" if settled_count > 0 else "没有新的记录需要结账！\n"
            if already_settled_count > 0:
                message += f"{already_settled_count} 条记录已结账，未重复处理。\n"
            message += f"结账日期: {settle_date}"
            tk.messagebox.showinfo("结账", message)
        ttk.Button(date_dialog, text="确定", command=confirm_settle).pack(pady=10)
    tree.bind("<Double-1>", on_settle_selected)

    # 鼠标悬停提示“双击结账”
    settle_tip = tk.Toplevel(tree)
    settle_tip.withdraw()
    settle_tip.overrideredirect(True)
    settle_label = tk.Label(settle_tip, text="双击结账", background="#ffffe0", relief=tk.SOLID, borderwidth=1, font=("微软雅黑", 9))
    settle_label.pack(ipadx=4)
    def show_settle_tip(event):
        row_id = tree.identify_row(event.y)
        if row_id:
            x = tree.winfo_rootx() + event.x + 30
            y = tree.winfo_rooty() + event.y + 20
            settle_tip.geometry(f"+{x}+{y}")
            settle_tip.deiconify()
        else:
            settle_tip.withdraw()
    def hide_settle_tip(event):
        settle_tip.withdraw()
    tree.bind('<Motion>', show_settle_tip)
    tree.bind('<Leave>', hide_settle_tip)
    for col, text in headers:
        tree.heading(col, text=text)
        if col == "in_date":
            tree.column(col, anchor=tk.CENTER, width=120)
        else:
            tree.column(col, anchor=tk.CENTER, width=80)
    def load_stock_data():
        for row in tree.get_children():
            tree.delete(row)
        for idx, row in enumerate(dbutil.get_all_stock(), 1):
            id_, factory, product_no, size, color,unit, in_quantity, price, total, is_settled, in_date = row
            is_settled = "是" if is_settled else "否"
            tree.insert("", tk.END, iid=str(id_), values=[idx, factory, product_no, size, color,unit, in_quantity, price, total, is_settled, in_date])
    load_stock_data()
    # 新增库存后刷新表格
    def open_add_stock_dialog_and_refresh():
        open_add_stock_dialog()
        load_stock_data()
    # 新增库存弹窗
    def open_add_stock_dialog():
        dialog = tk.Toplevel(main_win)
        dialog.title("新增库存")
        dialog.transient(main_win)
        center_window(dialog, 330, 200)
        fields = [
            ("厂家", "factory"),
            ("货号", "product_no"),
            ("尺码", "size"),
            ("颜色", "color"),
            ("单位", "unit"),
            ("入库数量", "in_quantity"),
            ("单价", "price"),
            ("合计", "total"),
            ("入库时间", "in_date")
        ]
        vars = {}
        entry_refs = {}
        for idx, (label, key) in enumerate(fields):
            if key == "in_date":
                vars[key] = tk.StringVar(value=utils.get_current_date())
            else:
                vars[key] = tk.StringVar()
        # 第一排：厂家、货号
        row1 = ttk.Frame(dialog)
        row1.pack(fill=tk.X, padx=10, pady=4, anchor=tk.W)
        ttk.Label(row1, text="厂家:").pack(side=tk.LEFT)
        entry_refs['factory'] = ttk.Entry(row1, textvariable=vars['factory'], width=17)
        entry_refs['factory'].pack(side=tk.LEFT, padx=(2,10))
        ttk.Label(row1, text="货号:").pack(side=tk.LEFT)
        entry_refs['product_no'] = ttk.Entry(row1, textvariable=vars['product_no'], width=17)
        entry_refs['product_no'].pack(side=tk.LEFT, padx=(2,0))

        # 第二排：尺码、颜色、单位
        row2 = ttk.Frame(dialog)
        row2.pack(fill=tk.X, padx=10, pady=4, anchor=tk.W)
        ttk.Label(row2, text="尺码:").pack(side=tk.LEFT)
        entry_refs['size'] = ttk.Entry(row2, textvariable=vars['size'], width=9)
        entry_refs['size'].pack(side=tk.LEFT, padx=(2,10))
        ttk.Label(row2, text="颜色:").pack(side=tk.LEFT)
        entry_refs['color'] = ttk.Entry(row2, textvariable=vars['color'], width=9)
        entry_refs['color'].pack(side=tk.LEFT, padx=(2,10))
        ttk.Label(row2, text="单位:").pack(side=tk.LEFT)
        vars['unit'] = tk.StringVar()
        entry_refs['unit'] = ttk.Entry(row2, textvariable=vars['unit'], width=7)
        entry_refs['unit'].pack(side=tk.LEFT, padx=(2,0))

        # 第三排：入库数量、单价
        row3 = ttk.Frame(dialog)
        row3.pack(fill=tk.X, padx=10, pady=4, anchor=tk.W)
        ttk.Label(row3, text="入库数量:").pack(side=tk.LEFT)
        entry_refs['in_quantity'] = ttk.Entry(row3, textvariable=vars['in_quantity'], width=15)
        entry_refs['in_quantity'].pack(side=tk.LEFT, padx=(2,10))
        vars['in_quantity'].trace_add('write', lambda *args: update_total(vars))
        ttk.Label(row3, text="单价:").pack(side=tk.LEFT)
        def validate_price(new_value):
            if new_value == "":
                return True
            if new_value.count('.') > 1:
                return False
            if new_value.startswith('.'):
                return False
            return all(c.isdigit() or c == '.' for c in new_value)
        vcmd = dialog.register(validate_price)
        entry_refs['price'] = ttk.Entry(row3, textvariable=vars['price'], validate="key", validatecommand=(vcmd, '%P'), width=15)
        entry_refs['price'].pack(side=tk.LEFT, padx=(2,0))
        vars['price'].trace_add('write', lambda *args: update_total(vars))

        # 第四排：合计、入库时间
        row4 = ttk.Frame(dialog)
        row4.pack(fill=tk.X, padx=10, pady=4, anchor=tk.W)
        ttk.Label(row4, text="合计:").pack(side=tk.LEFT)
        total_var = tk.StringVar()
        entry_refs['total'] = ttk.Entry(row4, textvariable=total_var, state='readonly', justify='right', width=15)
        entry_refs['total'].pack(side=tk.LEFT, padx=(2,10))
        vars['total'] = total_var
        ttk.Label(row4, text="入库时间:").pack(side=tk.LEFT)
        entry_refs['in_date'] = ttk.Entry(row4, textvariable=vars['in_date'], width=15)
        entry_refs['in_date'].pack(side=tk.LEFT, padx=(2,0))

        # 第五排：确定按钮
        def handle_stock_submit(is_edit=False, stock_id=None):
            for k, v in vars.items():
                if k != 'total' and not v.get().strip():
                    error_label['text'] = "所有字段不能为空！"
                    return
            try:
                vars["total"].set(utils.calculate_total(vars["in_quantity"].get(), vars["price"].get()))
                factory = vars['factory'].get().strip()
                product_no = vars['product_no'].get().strip()
                size = vars['size'].get().strip()
                color = vars['color'].get().strip()
                unit = vars['unit'].get().strip()
                in_quantity = int(vars['in_quantity'].get())
                price = float(vars['price'].get())
                total = float(vars['total'].get())
                in_date = vars['in_date'].get().strip()
                if is_edit:
                    dbutil.update_stock_by_id(
                        stock_id,
                        factory,
                        product_no,
                        size,
                        color,
                        unit,
                        in_quantity,
                        price,
                        total,
                        unit
                    )
                    dbutil.update_inventory_by_stock_id(
                        stock_id,
                        factory,
                        product_no,
                        size,
                        color,
                        unit,
                        in_quantity
                    )
                    tk.messagebox.showinfo("成功", "修改成功！")
                    dialog.destroy()  # 关闭对话框
                else:
                    dbutil.insert_stock(
                        factory,
                        product_no,
                        size,
                        color,
                        unit,
                        in_quantity,
                        price,
                        total,
                        in_date
                    )
                    stock_rows = dbutil.get_all_stock()
                    if stock_rows:
                        stock_id = stock_rows[0][0]
                    # 处理库存合并和日志
                    found = None
                    inventory_rows = dbutil.get_all_inventory()
                    for row in inventory_rows:
                        if row[2] == factory and row[3] == product_no and row[5] == color:
                            found = row
                            break
                    if found:
                        inventory_id = found[0]
                        old_size = found[4] or ''
                        new_size = size or ''
                        merged_size = dbutil.merge_size(old_size, new_size) if hasattr(dbutil, 'merge_size') else ','.join(sorted(set(s.strip() for s in (old_size + ',' + new_size).split(',') if s.strip())))
                        dbutil.update_inventory_size_by_id(inventory_id, merged_size)
                        dbutil.decrease_inventory_by_id(inventory_id, -in_quantity)
                        dbutil.insert_stock_log(
                            factory,
                            product_no,
                            merged_size,
                            color,
                            in_quantity,
                            '入库',
                            in_date
                        )
                        if hasattr(main_win, 'refresh_logs'):
                            main_win.refresh_logs()
                        tk.messagebox.showinfo("成功", f"已存在相同库存（尺码已合并为：{merged_size}），数量已增加 {in_quantity}！")
                    else:
                        dbutil.insert_inventory_from_stock(
                            stock_id,
                            factory,
                            product_no,
                            size,
                            color,
                            unit,
                            in_quantity
                        )
                        dbutil.insert_stock_log(
                            factory,
                            product_no,
                            size,
                            color,
                            in_quantity,
                            '入库',
                            in_date
                        )
                        if hasattr(main_win, 'refresh_logs'):
                            main_win.refresh_logs()
                        tk.messagebox.showinfo("成功", "新增成功！")
                    vars['color'].set("")
                    vars['in_quantity'].set("")
                    vars['price'].set("")
                    vars['total'].set("")
                    entry_refs['color'].focus_set()
                load_stock_data()
            except Exception as e:
                error_label['text'] = str(e)
        row5 = ttk.Frame(dialog)
        row5.pack(fill=tk.X, padx=10, pady=8, anchor=tk.W)
        add_btn = ttk.Button(row5, text="确定新增", command=lambda: handle_stock_submit(False), width=12)
        add_btn.pack(side=tk.LEFT)
        def update_total(vars):
            vars["total"].set(utils.calculate_total(vars["in_quantity"].get(), vars["price"].get()))
        dialog.bind('<Return>', lambda event: handle_stock_submit(False))
        # 创建样式来设置错误标签的前景色
        style = ttk.Style()
        style.configure("Error.TLabel", foreground="red")
        error_label = ttk.Label(dialog, text="", style="Error.TLabel", font=("微软雅黑", 10))
        error_label.pack(pady=(0, 5), padx=10, anchor=tk.W)
        dialog.after(100, lambda: entry_refs['factory'].focus_set())
        dialog.wait_window()
    # 删除/返厂和修改库存按钮已在创建时绑定，无需后续循环绑定
    def on_delete_or_return():
        selected = tree.selection()
        if not selected:
            tk.messagebox.showwarning("提示", "请先选择要操作的库存记录！")
            return
        item = tree.item(selected[0])
        values = item['values']
        factory, product_no, size, color = values[1], values[2], values[3], values[4]
        # 自定义操作对话框
        dialog = tk.Toplevel(main_win)
        dialog.title("请选择操作")
        dialog.transient(main_win)
        dialog.grab_set()
        center_window(dialog, 260, 120)
        msg = f"对【{factory} {product_no} {color} {size}】请选择操作："
        ttk.Label(dialog, text=msg, wraplength=240, font=("微软雅黑", 10)).pack(pady=(18, 10))
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=5)
        def do_return():
            dialog.destroy()
            stock_id = tree.selection()[0]
            return_qty = int(values[6])
            # 写入返厂日志
            dbutil.insert_stock_log(
                factory, product_no, size, color, return_qty, '返厂', utils.get_current_date()
            )
            # 返厂时，优先扣减所有厂家、货号、颜色一致的库存数量
            inventory_rows = dbutil.get_all_inventory()
            remain = return_qty
            for row in inventory_rows:
                # row: id, stock_id, factory, product_no, size, color, quantity
                if row[2] == factory and row[3] == product_no and row[5] == color:
                    inv_id = row[0]
                    inv_qty = row[7]
                    if inv_qty >= remain:
                        dbutil.decrease_inventory_by_id(inv_id, remain)
                        remain = 0
                        break
                    else:
                        dbutil.decrease_inventory_by_id(inv_id, inv_qty)
                        remain -= inv_qty
            # 同步删除库存表中对应记录
            dbutil.delete_inventory_by_stock_id(stock_id)
            dbutil.delete_stock_by_id(stock_id)
            # 自动刷新日志页面
            if hasattr(main_win, 'refresh_logs'):
                main_win.refresh_logs()
            load_stock_data()
            tk.messagebox.showinfo("返厂", f"已将【{factory} {product_no} {color} {size}】标记为返厂，并同步扣减库存！")
        def do_delete():
            dialog.destroy()
            stock_id = tree.selection()[0]
            del_qty = int(values[6])
            # 删除时，优先扣减所有厂家、货号、颜色一致的库存数量
            inventory_rows = dbutil.get_all_inventory()
            remain = del_qty
            for row in inventory_rows:
                if row[2] == factory and row[3] == product_no and row[5] == color:
                    inv_id = row[0]
                    inv_qty = row[7]
                    if inv_qty >= remain:
                        dbutil.decrease_inventory_by_id(inv_id, remain)
                        remain = 0
                        break
                    else:
                        dbutil.decrease_inventory_by_id(inv_id, inv_qty)
                        remain -= inv_qty
            # 先删除库存表中对应记录
            dbutil.delete_inventory_by_stock_id(stock_id)
            dbutil.delete_stock_by_id(stock_id)
            load_stock_data()
            tk.messagebox.showinfo("删除", "删除成功，库存已同步扣减！")
        def do_cancel():
            dialog.destroy()
        ttk.Button(btn_frame, text="返厂", width=8, command=do_return).pack(side=tk.LEFT, padx=8)
        ttk.Button(btn_frame, text="删除", width=8, command=do_delete).pack(side=tk.LEFT, padx=8)
        ttk.Button(btn_frame, text="取消", width=8, command=do_cancel).pack(side=tk.LEFT, padx=8)
        dialog.wait_window()
    # 修改库存弹窗
    def open_edit_stock_dialog():
        selected = tree.selection()
        if not selected:
            tk.messagebox.showwarning("提示", "请先选择要修改的库存记录！")
            return
        item = tree.item(selected[0])
        values = item['values']
        stock_id = selected[0]
        dialog = tk.Toplevel(main_win)
        dialog.title("修改入库")
        dialog.transient(main_win)
        dialog.grab_set()
        center_window(dialog, 330, 200)
        fields = [
            ("厂家", "factory"),
            ("货号", "product_no"),
            ("尺码", "size"),
            ("颜色", "color"),
            ("单位", "unit"),
            ("入库数量", "in_quantity"),
            ("单价", "price"),
            ("合计", "total"),
            ("入库时间", "in_date")  # 补充入库时间字段，防止KeyError
        ]
        vars = {}
        entry_refs = {}
        for idx, (label, key) in enumerate(fields):
            if key == "in_date":
                vars[key] = tk.StringVar(value=utils.get_current_date())
            else:
                vars[key] = tk.StringVar()
            if key in ["in_quantity", "price"]:
                vars[key].trace_add("write", lambda *args: vars["total"].set(utils.calculate_total(vars["in_quantity"].get(), vars["price"].get())))
        # 第一排：厂家、货号
        row1 = ttk.Frame(dialog)
        row1.pack(fill=tk.X, padx=10, pady=4, anchor=tk.W)
        ttk.Label(row1, text="厂家:").pack(side=tk.LEFT)
        entry_refs['factory'] = ttk.Entry(row1, textvariable=vars['factory'], width=17)
        entry_refs['factory'].pack(side=tk.LEFT, padx=(2,10))
        ttk.Label(row1, text="货号:").pack(side=tk.LEFT)
        entry_refs['product_no'] = ttk.Entry(row1, textvariable=vars['product_no'], width=17)
        entry_refs['product_no'].pack(side=tk.LEFT, padx=(2,0))

        # 第二排：尺码、颜色、单位
        row2 = ttk.Frame(dialog)
        row2.pack(fill=tk.X, padx=10, pady=4, anchor=tk.W)
        ttk.Label(row2, text="尺码:").pack(side=tk.LEFT)
        entry_refs['size'] = ttk.Entry(row2, textvariable=vars['size'], width=9)
        entry_refs['size'].pack(side=tk.LEFT, padx=(2,10))
        ttk.Label(row2, text="颜色:").pack(side=tk.LEFT)
        entry_refs['color'] = ttk.Entry(row2, textvariable=vars['color'], width=9)
        entry_refs['color'].pack(side=tk.LEFT, padx=(2,10))
        ttk.Label(row2, text="单位:").pack(side=tk.LEFT)
        vars['unit'] = tk.StringVar()
        entry_refs['unit'] = ttk.Entry(row2, textvariable=vars['unit'], width=7)
        entry_refs['unit'].pack(side=tk.LEFT, padx=(2,0))

        # 第三排：入库数量、单价
        row3 = ttk.Frame(dialog)
        row3.pack(fill=tk.X, padx=10, pady=4, anchor=tk.W)
        ttk.Label(row3, text="入库数量:").pack(side=tk.LEFT)
        entry_refs['in_quantity'] = ttk.Entry(row3, textvariable=vars['in_quantity'], width=15)
        entry_refs['in_quantity'].pack(side=tk.LEFT, padx=(2,10))
        vars['in_quantity'].trace_add('write', lambda *args: update_total(vars))
        ttk.Label(row3, text="单价:").pack(side=tk.LEFT)
        def validate_price(new_value):
            if new_value == "":
                return True
            if new_value.count('.') > 1:
                return False
            if new_value.startswith('.'):
                return False
            return all(c.isdigit() or c == '.' for c in new_value)
        vcmd = dialog.register(validate_price)
        entry_refs['price'] = ttk.Entry(row3, textvariable=vars['price'], validate="key", validatecommand=(vcmd, '%P'), width=15)
        entry_refs['price'].pack(side=tk.LEFT, padx=(2,0))
        vars['price'].trace_add('write', lambda *args: update_total(vars))

        # 第四排：合计、入库时间
        row4 = ttk.Frame(dialog)
        row4.pack(fill=tk.X, padx=10, pady=4, anchor=tk.W)
        ttk.Label(row4, text="合计:").pack(side=tk.LEFT)
        total_var = tk.StringVar()
        entry_refs['total'] = ttk.Entry(row4, textvariable=total_var, state='readonly', justify='right', width=15)
        entry_refs['total'].pack(side=tk.LEFT, padx=(2,10))
        vars['total'] = total_var
        ttk.Label(row4, text="入库时间:").pack(side=tk.LEFT)
        entry_refs['in_date'] = ttk.Entry(row4, textvariable=vars['in_date'], width=15)
        entry_refs['in_date'].pack(side=tk.LEFT, padx=(2,0))

        # 第五排：确定按钮
        row5 = ttk.Frame(dialog)
        row5.pack(fill=tk.X, padx=10, pady=8, anchor=tk.W)
        edit_btn = ttk.Button(row5, text="确定修改", command=lambda: handle_stock_submit(vars, entry_refs, error_label, dialog, True, stock_id), width=12)
        edit_btn.pack(side=tk.LEFT)
        def update_total(*args):
            vars['total'].set(utils.calculate_total(vars['in_quantity'].get(), vars['price'].get()))
        vars['in_quantity'].trace_add('write', update_total)
        vars['price'].trace_add('write', update_total)
        dialog.bind('<Return>', lambda event: handle_stock_submit(vars, entry_refs, error_label, dialog, True, stock_id))
        error_label = ttk.Label(dialog, text="", style="Error.TLabel", font=("微软雅黑", 10))
        error_label.pack(pady=(0, 5), padx=10, anchor=tk.W)
        dialog.after(100, lambda: entry_refs['factory'].focus_set())
        # 修正字段赋值顺序，确保各字段对应正确的 values 索引
        # values: [idx, factory, product_no, size, color, unit, in_quantity, price, total, is_settled, in_date]
        vars["factory"].set(values[1])
        vars["product_no"].set(values[2])
        vars["size"].set(values[3])
        vars["color"].set(values[4])
        vars["unit"].set(values[5])
        vars["in_quantity"].set(values[6])
        vars["price"].set(values[7])
        vars["total"].set(values[8])
        vars["in_date"].set(values[10])
    # 修改按钮已在创建时绑定，无需后续循环绑定
    # 恢复返厂（删除/返厂）按钮
    btn_del = ttk.Button(stock_toolbar, text="🗑️", width=3, command=on_delete_or_return)
    btn_del.pack(side=tk.LEFT, padx=3)
    add_tooltip(btn_del, "删除/返厂")

    # 恢复编辑（修改）按钮
    btn_edit = ttk.Button(stock_toolbar, text="✏️", width=3, command=open_edit_stock_dialog)
    btn_edit.pack(side=tk.LEFT, padx=3)
    add_tooltip(btn_edit, "修改库存")

    # 批量导入按钮（图标形式）
    btn_batch_import = ttk.Button(stock_toolbar, text="📥", width=3, command=handle_batch_import)
    btn_batch_import.pack(side=tk.LEFT, padx=3)
    add_tooltip(btn_batch_import, "批量导入库存")

    # 新增结账按钮（批量结账）
    btn_settle = ttk.Button(stock_toolbar, text="结账", width=6)
    btn_settle.pack(side=tk.LEFT, padx=3)
    add_tooltip(btn_settle, "批量结账")
    def batch_settle():
        selected = tree.selection()
        if not selected:
            tk.messagebox.showwarning("提示", "请先选择要结账的库存记录！")
            return
        # 创建日期选择对话框
        date_dialog = tk.Toplevel(main_win)
        date_dialog.title("选择结账日期")
        date_dialog.transient(main_win)
        date_dialog.grab_set()
        center_window(date_dialog, 300, 120)
        ttk.Label(date_dialog, text="请选择结账日期:").pack(pady=10)
        date_var = tk.StringVar()
        date_entry = ttk.Entry(date_dialog, textvariable=date_var)
        date_entry.pack(pady=5)
        date_entry.insert(0, utils.get_current_date())  # 默认当前日期
        def confirm_settle():
            settle_date = date_var.get()
            if not settle_date:
                tk.messagebox.showwarning("提示", "请选择结账日期！")
                return
            # 筛选未结账和已结账物品
            unsettled_items = []
            settled_items = []
            for iid in selected:
                item = tree.item(iid)
                values = item['values']
                if values[9] == "是":
                    settled_items.append(values)
                else:
                    unsettled_items.append((iid, item))
            
            # 只处理未结账物品
            settled_count = 0
            for iid, item in unsettled_items:
                dbutil.settle_stock_by_id(iid)
                # 写入结账日志
                values = item['values']
                dbutil.insert_settle_log(
                    values[1],  # factory
                    values[2],  # product_no
                    values[3],  # size
                    values[4],  # color
                    int(values[6]),  # in_quantity
                    float(values[7]),  # price
                    float(values[8]),  # total
                    settle_date
                )
                settled_count += 1
            
            # 自动刷新日志页面
            if hasattr(main_win, 'refresh_logs'):
                main_win.refresh_logs()
            load_stock_data()
            date_dialog.destroy()
            
            # 显示结账结果统计
            message = f"成功结账 {settled_count} 条记录！\n" if settled_count > 0 else "没有新的记录需要结账！\n"
            if len(settled_items) > 0:
                message += f"{len(settled_items)} 条记录已结账，未重复处理。\n"
            message += f"结账日期: {settle_date}"
            tk.messagebox.showinfo("结账", message)
        ttk.Button(date_dialog, text="确定", command=confirm_settle).pack(pady=10)
    btn_settle.config(command=batch_settle)
    return frame